#公共方法
import csv

from ddt import data
from openpyxl import load_workbook, workbook
from openpyxl.xml.constants import MIN_ROW



#读取csv文件
def read_csv():
    with (open("./file_data/requests_body.csv",'r',encoding='utf-8')) as f:

        csv_data = csv.reader(f)
        #print(csv_data)
        csv_data.__next__()
        
        data_list = []
        for request_data in csv_data:

            #print(request_data)
            data_list.append(request_data)
        print(data_list)

        return data_list
        

        # for row in range(len(data_list)):
        #     for col in range(len(data_list[row])):
        #         r = row + 1
        #         c = col + 1
        #         print (data_list[r][c])


#读取excl文件
def read_excl():

    wb = load_workbook(filename='./file_data/request_excl.xlsx')
    sheet = wb['Sheet1']
    #print (sheet['A1'].value)

    re_list = []

    for row in sheet.iter_rows(min_row = 2) :
        rows = []
        print (row)
        for real_data in row:
            rows.append(real_data.value)
        re_list.append(rows)    
    
    print (re_list)
    wb.close
 
    return re_list








if __name__ == "__main__":
    read_excl()